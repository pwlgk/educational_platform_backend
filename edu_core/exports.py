import datetime
import logging
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from django.utils import timezone
from django.db.models import Prefetch, Q, Sum, F 
from django.http import HttpResponse
from users.models import User
from .models import (
    Lesson, LessonJournalEntry, StudentGroup, Subject, StudyPeriod, AcademicYear,
    Attendance, Grade, Homework, HomeworkSubmission
)
from django.utils.translation import gettext_lazy as _
from django.utils.encoding import force_str # <--- ДОБАВЛЕН ИМПОРТ

logger = logging.getLogger(__name__)

class JournalExporter:
    def __init__(self, request_user, filters=None):
        self.user = request_user
        self.filters = filters if filters else {}
        self.workbook = Workbook()
        self.workbook.remove(self.workbook.active) 
        self.applied_filters_text = self._get_applied_filters_text()

    def _get_applied_filters_text(self):
        filter_descs = []
        if self.filters.get('academic_year_id'):
            try:
                year = AcademicYear.objects.get(id=self.filters['academic_year_id'])
                filter_descs.append(force_str(_("Учебный год: %(year_name)s")) % {'year_name': year.name})
            except AcademicYear.DoesNotExist: pass
        if self.filters.get('study_period_id'):
            try:
                period = StudyPeriod.objects.get(id=self.filters['study_period_id'])
                filter_descs.append(force_str(_("Учебный период: %(period_name)s")) % {'period_name': period.name})
            except StudyPeriod.DoesNotExist: pass
        if self.filters.get('student_group_id'):
            try:
                group = StudentGroup.objects.get(id=self.filters['student_group_id'])
                filter_descs.append(force_str(_("Группа: %(group_name)s")) % {'group_name': group.name})
            except StudentGroup.DoesNotExist: pass
        if self.filters.get('subject_id'):
            try:
                subject = Subject.objects.get(id=self.filters['subject_id'])
                filter_descs.append(force_str(_("Предмет: %(subject_name)s")) % {'subject_name': subject.name})
            except Subject.DoesNotExist: pass
        if self.filters.get('teacher_id'): 
            try:
                teacher = User.objects.get(id=self.filters['teacher_id'], role=User.Role.TEACHER)
                filter_descs.append(force_str(_("Преподаватель: %(teacher_name)s")) % {'teacher_name': teacher.get_full_name()})
            except User.DoesNotExist: pass
        
        date_from = self.filters.get('date_from')
        date_to = self.filters.get('date_to')
        if date_from and date_to:
            filter_descs.append(force_str(_("Даты: с %(date_from)s по %(date_to)s")) % {'date_from': date_from.strftime('%d.%m.%Y'), 'date_to': date_to.strftime('%d.%m.%Y')})
        elif date_from:
            filter_descs.append(force_str(_("Даты: с %(date_from)s")) % {'date_from': date_from.strftime('%d.%m.%Y')})
        elif date_to:
            filter_descs.append(force_str(_("Даты: по %(date_to)s")) % {'date_to': date_to.strftime('%d.%m.%Y')})
            
        return "; ".join(filter_descs) if filter_descs else force_str(_("Все данные"))

    def _add_filter_header_to_sheet(self, ws, sheet_title_info=""):
        title = force_str(_("Журнал успеваемости и посещаемости"))
        if sheet_title_info:
            title += f" ({sheet_title_info})"
        
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
        ws['A1'] = title
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal="center")

        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=5)
        ws['A2'] = force_str(_("Фильтры: %(filters_text)s")) % {'filters_text': self.applied_filters_text}
        ws['A2'].font = Font(italic=True, size=10)
        ws['A2'].alignment = Alignment(horizontal="center")
        ws.append([]) 

    def _apply_common_styles(self, ws, header_row_index=4):
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        for cell in ws[header_row_index]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        for row in ws.iter_rows(min_row=header_row_index):
            for cell in row:
                cell.border = thin_border
                if cell.row > header_row_index: 
                    cell.alignment = Alignment(vertical="top", wrap_text=True)

        for col_idx, column_cells in enumerate(ws.columns):
            max_length = 0
            column_letter = get_column_letter(col_idx + 1)
            for cell in column_cells[header_row_index-1:]:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except: pass
            adjusted_width = (max_length + 2) * 1.1 
            ws.column_dimensions[column_letter].width = min(adjusted_width, 50)

        ws.freeze_panes = ws.cell(row=header_row_index + 1, column=1)

    def _get_base_lesson_queryset(self):
        return Lesson.objects.select_related(
            'study_period__academic_year', 
            'student_group__curator', 
            'subject', 
            'teacher', 
            'classroom'
        ).prefetch_related(
            Prefetch('journal_entry', queryset=LessonJournalEntry.objects.prefetch_related(
                Prefetch('attendances', queryset=Attendance.objects.select_related('student')),
                Prefetch('homework_assignments', queryset=Homework.objects.prefetch_related(
                    Prefetch('submissions', queryset=HomeworkSubmission.objects.select_related('student', 'grade_for_submission'))
                ))
            )),
            Prefetch('grades_for_lesson_instance', queryset=Grade.objects.filter(grade_type=Grade.GradeType.LESSON_WORK).select_related('student'))
        ).order_by('start_time', 'student_group__name')

    def _filter_queryset_by_params(self, queryset):
        if self.filters.get('academic_year_id'):
            queryset = queryset.filter(study_period__academic_year_id=self.filters['academic_year_id'])
        if self.filters.get('study_period_id'):
            queryset = queryset.filter(study_period_id=self.filters['study_period_id'])
        if self.filters.get('student_group_id'):
            queryset = queryset.filter(student_group_id=self.filters['student_group_id'])
        if self.filters.get('subject_id'):
            queryset = queryset.filter(subject_id=self.filters['subject_id'])
        if self.filters.get('teacher_id'):
            queryset = queryset.filter(teacher_id=self.filters['teacher_id'])
        
        date_from = self.filters.get('date_from')
        date_to = self.filters.get('date_to')
        if date_from: queryset = queryset.filter(start_time__date__gte=date_from)
        if date_to: queryset = queryset.filter(start_time__date__lte=date_to)
            
        return queryset.distinct()

    def export_teacher_journal(self):
        logger.info(f"Teacher {self.user.email} requested journal export with filters: {self.filters}")
        base_queryset = self._get_base_lesson_queryset()
        
        teacher_lessons_qs = base_queryset.filter(teacher=self.user)
        curated_group_ids = StudentGroup.objects.filter(curator=self.user).values_list('id', flat=True)
        curator_lessons_qs = base_queryset.filter(student_group_id__in=list(curated_group_ids))
        
        final_queryset = self._filter_queryset_by_params((teacher_lessons_qs | curator_lessons_qs).distinct())

        lessons_by_subject_group = {}
        for lesson in final_queryset:
            key = (lesson.subject.name, lesson.student_group.name, lesson.student_group_id, lesson.subject_id)
            if key not in lessons_by_subject_group: lessons_by_subject_group[key] = []
            lessons_by_subject_group[key].append(lesson)

        if not lessons_by_subject_group:
            ws = self.workbook.create_sheet(title=force_str(_("Нет данных")))
            self._add_filter_header_to_sheet(ws)
            ws.append([force_str(_("Нет данных для экспорта по указанным фильтрам."))])
            self._apply_common_styles(ws, header_row_index=4)
            logger.warning(f"No data found for teacher {self.user.email} export with filters: {self.filters}")
            return self._save_workbook_to_response(self._generate_filename("Teacher_Journal_NoData"))

        for (subject_name, group_name, group_id, subject_id_unused), lessons_in_group_subject in lessons_by_subject_group.items():
            sheet_title_info = f"{subject_name} - {group_name}"
            safe_sheet_title = "".join(c if c.isalnum() else "_" for c in sheet_title_info)[:31]
            ws = self.workbook.create_sheet(title=safe_sheet_title)
            self._add_filter_header_to_sheet(ws, sheet_title_info)

            student_data_headers_lazy = [
                '№', _('ФИО студента'), _('Присутствие'), _('Комм. к присут.'), 
                _('Оценка за урок'), _('Комм. к оценке'),
                _('Статус ДЗ'), _('Оценка за ДЗ'), _('Комм. к оценке ДЗ')
            ]
            student_data_headers = [force_str(h) if not isinstance(h, str) else h for h in student_data_headers_lazy]
            
            students_in_group = StudentGroup.objects.get(id=group_id).students.order_by('last_name', 'first_name')
            
            for lesson in sorted(lessons_in_group_subject, key=lambda l: l.start_time):
                ws.append([]) 
                lesson_info_row_idx = ws.max_row
                
                journal_entry = getattr(lesson, 'journal_entry', None)
                homework = None
                if journal_entry and hasattr(journal_entry, 'homework_assignments') and journal_entry.homework_assignments.exists():
                    homework = journal_entry.homework_assignments.first()

                lesson_info_str = (
                    f"{force_str(_('Занятие'))}: {lesson.start_time.strftime('%d.%m.%Y %H:%M')}-{lesson.end_time.strftime('%H:%M')}; "
                    f"{force_str(_('Тема'))}: {getattr(journal_entry, 'topic_covered', '-')}; "
                    f"{force_str(_('ДЗ'))}: {homework.title if homework else '-'}"
                )
                ws.cell(row=lesson_info_row_idx, column=1, value=lesson_info_str)
                ws.merge_cells(start_row=lesson_info_row_idx, start_column=1, end_row=lesson_info_row_idx, end_column=len(student_data_headers))
                ws.cell(row=lesson_info_row_idx, column=1).font = Font(bold=True)
                ws.cell(row=lesson_info_row_idx, column=1).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

                ws.append(student_data_headers)
                current_header_row_for_style = ws.max_row
                for cell in ws[current_header_row_for_style]:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

                for student_idx, student in enumerate(students_in_group):
                    attendance_obj, lesson_grade_obj, hw_submission_obj, hw_grade_obj, status_dz = None, None, None, None, force_str(_("Нет ДЗ"))
                    if journal_entry and hasattr(journal_entry, 'attendances'):
                        for att in journal_entry.attendances.all():
                            if att.student_id == student.id: attendance_obj = att; break
                    if hasattr(lesson, 'grades_for_lesson_instance'):
                        for grade in lesson.grades_for_lesson_instance.all():
                             if grade.student_id == student.id: lesson_grade_obj = grade; break
                    if homework:
                        status_dz = force_str(_("Не сдано"))
                        if hasattr(homework, 'submissions'):
                            for sub in homework.submissions.all():
                                if sub.student_id == student.id:
                                    hw_submission_obj = sub; status_dz = force_str(_("Сдано (ожидает)"))
                                    if hasattr(sub, 'grade_for_submission') and sub.grade_for_submission:
                                        hw_grade_obj = sub.grade_for_submission; status_dz = force_str(_("Оценено: %(grade_val)s")) % {'grade_val': hw_grade_obj.grade_value}
                                    break
                        if not hw_submission_obj and homework.due_date and timezone.now().date() > homework.due_date.date():
                             status_dz = force_str(_("Не сдано (просрочено)"))
                    row_data = [
                        student_idx + 1, student.get_full_name(),
                        attendance_obj.get_status_display() if attendance_obj else '-',
                        attendance_obj.comment if attendance_obj and attendance_obj.comment else '-',
                        lesson_grade_obj.grade_value if lesson_grade_obj else '-',
                        lesson_grade_obj.comment if lesson_grade_obj and lesson_grade_obj.comment else '-',
                        status_dz,
                        hw_grade_obj.grade_value if hw_grade_obj else '-',
                        hw_grade_obj.comment if hw_grade_obj and hw_grade_obj.comment else '-',
                    ]
                    ws.append(row_data)
            
            self._apply_common_styles(ws, header_row_index=3) 
            logger.info(f"Generated sheet '{safe_sheet_title}' for teacher {self.user.email}")

        return self._save_workbook_to_response(self._generate_filename("Teacher_Journal"))

    def export_admin_journal(self):
        logger.info(f"Admin {self.user.email} requested journal export with filters: {self.filters}")
        base_queryset = self._get_base_lesson_queryset()
        final_queryset = self._filter_queryset_by_params(base_queryset)

        if not final_queryset.exists():
            ws = self.workbook.create_sheet(title=force_str(_("Нет данных")))
            self._add_filter_header_to_sheet(ws)
            ws.append([force_str(_("Нет данных для экспорта по указанным фильтрам."))])
            self._apply_common_styles(ws, header_row_index=4)
            logger.warning(f"No data found for admin export with filters: {self.filters}")
            return self._save_workbook_to_response(self._generate_filename("Admin_Journal_NoData"))

        ws = self.workbook.create_sheet(title=force_str(_("Общий журнал")))
        self._add_filter_header_to_sheet(ws)
        
        headers_lazy = [
            _('ID Занятия'), _('Дата'), _('Начало'), _('Конец'), _('Уч. Год'), _('Уч. Период'),
            _('Группа'), _('Предмет'), _('Тип занятия'), _('Преподаватель'), _('Аудитория'), _('Тема урока'),
            _('ФИО студента'), _('ID Студента'), _('Присутствие'), _('Комм. присут.'), 
            _('Оценка (урок)'), _('Комм. (урок)'), _('ДЗ ID'), _('ДЗ описание'), 
            _('Статус ДЗ'), _('Оценка (ДЗ)'), _('Комм. (ДЗ)')
        ]
        headers = [force_str(h) for h in headers_lazy]
        ws.append(headers)

        for lesson in final_queryset:
            journal_entry = getattr(lesson, 'journal_entry', None)
            homework = None
            if journal_entry and hasattr(journal_entry, 'homework_assignments') and journal_entry.homework_assignments.exists():
                 homework = journal_entry.homework_assignments.first()

            for student in lesson.student_group.students.order_by('last_name', 'first_name'):
                attendance_obj, lesson_grade_obj, hw_submission_obj, hw_grade_obj, status_dz = None, None, None, None, force_str(_("Нет ДЗ"))
                if journal_entry and hasattr(journal_entry, 'attendances'):
                    for att in journal_entry.attendances.all():
                        if att.student_id == student.id: attendance_obj = att; break
                if hasattr(lesson, 'grades_for_lesson_instance'):
                    for grade in lesson.grades_for_lesson_instance.all():
                         if grade.student_id == student.id: lesson_grade_obj = grade; break
                if homework:
                    status_dz = force_str(_("Не сдано"))
                    if hasattr(homework, 'submissions'):
                        for sub in homework.submissions.all():
                            if sub.student_id == student.id:
                                hw_submission_obj = sub; status_dz = force_str(_("Сдано (ожидает)"))
                                if hasattr(sub, 'grade_for_submission') and sub.grade_for_submission:
                                    hw_grade_obj = sub.grade_for_submission; status_dz = force_str(_("Оценено: %(grade_val)s")) % {'grade_val': hw_grade_obj.grade_value}
                                break
                    if not hw_submission_obj and homework.due_date and timezone.now().date() > homework.due_date.date():
                         status_dz = force_str(_("Не сдано (просрочено)"))
                row_data = [
                    lesson.id, lesson.start_time.strftime('%d.%m.%Y'), lesson.start_time.strftime('%H:%M'), lesson.end_time.strftime('%H:%M'),
                    lesson.study_period.academic_year.name, lesson.study_period.name,
                    lesson.student_group.name, lesson.subject.name, lesson.get_lesson_type_display(),
                    lesson.teacher.get_full_name() if lesson.teacher else '-',
                    lesson.classroom.identifier if lesson.classroom else '-',
                    getattr(journal_entry, 'topic_covered', '-'),
                    student.get_full_name(), student.id,
                    attendance_obj.get_status_display() if attendance_obj else '-',
                    attendance_obj.comment if attendance_obj and attendance_obj.comment else '-',
                    lesson_grade_obj.grade_value if lesson_grade_obj else '-',
                    lesson_grade_obj.comment if lesson_grade_obj and lesson_grade_obj.comment else '-',
                    homework.id if homework else '-',
                    homework.description if homework and homework.description else '-',
                    status_dz,
                    hw_grade_obj.grade_value if hw_grade_obj else '-',
                    hw_grade_obj.comment if hw_grade_obj and hw_grade_obj.comment else '-',
                ]
                ws.append(row_data)
        
        self._apply_common_styles(ws, header_row_index=4)
        logger.info(f"Generated admin journal sheet for user {self.user.email}")
        return self._save_workbook_to_response(self._generate_filename("Admin_Full_Journal"))

    def _generate_filename(self, base_name="Journal_Export"):
        filename_parts = [base_name]
        try:
            if self.filters.get('academic_year_id'):
                year = AcademicYear.objects.get(id=self.filters['academic_year_id'])
                filename_parts.append(str(year.name).replace(" ", "_").replace("/", "-"))
            if self.filters.get('study_period_id'):
                period = StudyPeriod.objects.get(id=self.filters['study_period_id'])
                filename_parts.append(str(period.name).replace(" ", "_"))
            if self.filters.get('student_group_id'):
                group = StudentGroup.objects.get(id=self.filters['student_group_id'])
                filename_parts.append(str(group.name).replace(" ", "_"))
            if self.filters.get('subject_id'):
                subject = Subject.objects.get(id=self.filters['subject_id'])
                filename_parts.append(str(subject.name).replace(" ", "_"))
            if self.filters.get('teacher_id'):
                teacher = User.objects.get(id=self.filters['teacher_id'])
                filename_parts.append(f"Teacher_{teacher.last_name}")
            
            date_from = self.filters.get('date_from')
            date_to = self.filters.get('date_to')
            if date_from: filename_parts.append(f"from_{date_from.strftime('%Y%m%d')}")
            if date_to: filename_parts.append(f"to_{date_to.strftime('%Y%m%d')}")
        except Exception as e:
            logger.warning(f"Could not retrieve some filter names for filename generation: {e}")
        
        return "_".join(filename_parts) + ".xlsx"

    def _save_workbook_to_response(self, filename="journal_export.xlsx"):
        excel_io = BytesIO()
        self.workbook.save(excel_io)
        excel_io.seek(0)
        response = HttpResponse(
            excel_io.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        logger.info(f"Prepared HttpResponse with Excel file: {filename}")
        return response